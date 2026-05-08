from __future__ import annotations

import csv
import io
import os
import sqlite3
import string
import tempfile
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable
from zipfile import ZipFile
from xml.etree import ElementTree as ET

from django.conf import settings
from django.db import connections
from openpyxl import Workbook, load_workbook

ABJAD_MAP = {
    "ا": 1,
    "ب": 2,
    "ج": 3,
    "د": 4,
    "ه": 5,
    "و": 6,
    "ز": 7,
    "ح": 8,
    "ط": 9,
    "ی": 10,
    "ک": 20,
    "ل": 30,
    "م": 40,
    "ن": 50,
    "س": 60,
    "ع": 70,
    "ف": 80,
    "ص": 90,
    "ق": 100,
    "ر": 200,
    "ش": 300,
    "ت": 400,
    "ث": 500,
    "خ": 600,
    "ذ": 700,
    "ض": 800,
    "ظ": 900,
    "غ": 1000,
    "آ": 1,
    "ء": 1,
    "ئ": 1,
    "ة": 400,
    "أ": 1,
    "ؤ": 6,
}
ABJAD_SORT_INDEX = {char: index for index, char in enumerate(ABJAD_MAP.keys())}

DOT_MAP = {
    "ا": 0,
    "ب": 1,
    "ج": 1,
    "د": 0,
    "ه": 0,
    "و": 0,
    "ز": 1,
    "ح": 0,
    "ط": 0,
    "ی": 2,
    "ک": 0,
    "ل": 0,
    "م": 0,
    "ن": 1,
    "س": 0,
    "ع": 0,
    "ف": 1,
    "ص": 0,
    "ق": 2,
    "ر": 0,
    "ش": 3,
    "ت": 2,
    "ث": 3,
    "خ": 1,
    "ذ": 1,
    "ض": 1,
    "ظ": 1,
    "غ": 1,
    "آ": 0,
    "ء": 0,
    "ئ": 0,
    "ة": 2,
    "أ": 0,
    "ؤ": 0,
}

PRONOUNCED_MAP = {
    "ا": 111,
    "ب": 4,
    "ج": 53,
    "د": 35,
    "ه": 7,
    "و": 13,
    "ز": 9,
    "ح": 10,
    "ط": 11,
    "ی": 12,
    "ک": 101,
    "ل": 71,
    "م": 90,
    "ن": 106,
    "س": 120,
    "ع": 130,
    "ف": 82,
    "ص": 95,
    "ق": 181,
    "ر": 202,
    "ش": 360,
    "ت": 402,
    "ث": 502,
    "خ": 602,
    "ذ": 731,
    "ض": 805,
    "ظ": 902,
    "غ": 1060,
    "آ": 111,
    "ء": 111,
    "ئ": 111,
    "ة": 402,
    "أ": 111,
    "ؤ": 13,
}

ALIF_MAP = {"ا": 1, "أ": 1, "آ": 1}

SAGHIR_MAP = {
    "ا": 1,
    "ب": 2,
    "ج": 3,
    "د": 4,
    "ه": 5,
    "و": 6,
    "ز": 7,
    "ح": 8,
    "ط": 9,
    "ی": 10,
    "ک": 8,
    "ل": 6,
    "م": 4,
    "ن": 2,
    "س": 0,
    "ع": 10,
    "ف": 8,
    "ص": 6,
    "ق": 4,
    "ر": 8,
    "ش": 0,
    "ت": 4,
    "ث": 8,
    "خ": 0,
    "ذ": 4,
    "ض": 8,
    "ظ": 0,
    "غ": 4,
    "آ": 1,
    "ء": 1,
    "ئ": 1,
    "ة": 4,
    "أ": 1,
    "ؤ": 6,
}

SORTABLE_COLUMNS = {
    "row_number": "ردیف",
    "phrase": "عبارت",
    "abjad_value": "عدد ابجد",
    "prime_index": "چندمین عدد اول",
    "digit_root": "ریشه عدد",
    "abjad_sum": "مجموع عدد ابجد",
    "parity_label": "زوج یا فرد",
    "parity_order": "چندمین زوج فرد",
    "letter_count": "تعداد حروف",
    "dot_count": "تعداد نقطه",
    "unique_letter_count": "تعداد حروف یکتا",
    "used_letters": "حروف استفاده شده",
    "pronounced_value": "عدد ملفوظی",
    "alif_count": "تعداد الف",
    "abjad_saghir": "ابجد صغیر",
}

VISIBLE_COLUMNS = [
    "row_number",
    "phrase",
    "abjad_value",
    "prime_index",
    "digit_root",
    "abjad_sum",
    "parity_label",
    "parity_order",
    "letter_count",
    "dot_count",
    "unique_letter_count",
    "used_letters",
    "pronounced_value",
    "alif_count",
    "abjad_saghir",
    "breakdown",
]

EXPORT_HEADERS = {
    "row_number": "ردیف",
    "phrase": "عبارت",
    "abjad_value": "عدد ابجد",
    "prime_index": "چندمین عدد اول",
    "digit_root": "ریشه عدد",
    "abjad_sum": "مجموع عدد ابجد",
    "parity_label": "زوج یا فرد",
    "parity_order": "چندمین زوج فرد",
    "letter_count": "تعداد حروف",
    "dot_count": "تعداد نقطه",
    "unique_letter_count": "تعداد حروف یکتا",
    "used_letters": "حروف استفاده شده",
    "pronounced_value": "عدد ملفوظی",
    "alif_count": "تعداد الف",
    "abjad_saghir": "ابجد صغیر",
    "breakdown": "تبدیل حرف به عدد",
}

PERSIAN_NORMALIZATION = str.maketrans(
    {
        "ي": "ی",
        "ى": "ی",
        "ك": "ک",
        "\u200c": " ",
        "\u200f": "",
        "\u200e": "",
        "\ufeff": "",
    }
)

ALLOWED_PUNCTUATION_CHARS = set(string.punctuation + "،؛؟«»ـ…–—“”‘’")
DUPLICATE_IGNORED_CHARS = set(string.whitespace).union(ALLOWED_PUNCTUATION_CHARS)
ALLOWED_EXTRA_PHRASE_CHARS = set(string.whitespace).union(ALLOWED_PUNCTUATION_CHARS)

PRIME_CACHE: list[int] = [2]


@dataclass(slots=True)
class CalculationResult:
    phrase: str
    normalized_phrase: str
    abjad_value: int
    prime_index: int
    digit_root: int
    abjad_sum: int
    parity_label: str
    parity_order: int
    letter_count: int
    dot_count: int
    unique_letter_count: int
    used_letters: str
    pronounced_value: int
    alif_count: int
    abjad_saghir: int
    breakdown: str

    def as_model_data(self) -> dict:
        return asdict(self)


def normalize_phrase(phrase: str) -> str:
    normalized = (phrase or "").translate(PERSIAN_NORMALIZATION)
    normalized = " ".join(normalized.split())
    return normalized.strip()


def duplicate_key(phrase: str) -> str:
    normalized = normalize_phrase(phrase)
    return "".join(char for char in normalized if char not in DUPLICATE_IGNORED_CHARS)


def find_invalid_phrase_chars(phrase: str) -> list[str]:
    normalized = normalize_phrase(phrase)
    invalid = []
    for char in normalized:
        if char in ABJAD_MAP or char in ALLOWED_EXTRA_PHRASE_CHARS:
            continue
        if char not in invalid:
            invalid.append(char)
    return invalid


def iter_mapped_values(phrase: str, mapping: dict[str, int]) -> list[int]:
    return [mapping[char] for char in phrase if char in mapping]


def compute_used_letters(phrase: str) -> tuple[int, str]:
    unique_letters = {char for char in phrase if char in ABJAD_MAP}
    ordered_letters = sorted(unique_letters, key=lambda char: (-ABJAD_MAP[char], ABJAD_SORT_INDEX.get(char, 9999), char))
    return len(ordered_letters), " - ".join(ordered_letters)


def compute_dot_count(phrase: str) -> int:
    total = 0
    for index, char in enumerate(phrase):
        if char not in DOT_MAP:
            continue
        if char == "ی":
            is_last = index == len(phrase) - 1
            next_is_space = index < len(phrase) - 1 and phrase[index + 1] == " "
            if is_last or next_is_space:
                continue
        total += DOT_MAP[char]
    return total


def compute_digit_root(number: int) -> int:
    value = number
    while value >= 10:
        value = sum(int(char) for char in str(value))
    return value


def is_prime(number: int) -> bool:
    if number < 2:
        return False
    if number == 2:
        return True
    if number % 2 == 0:
        return False
    divisor = 3
    while divisor * divisor <= number:
        if number % divisor == 0:
            return False
        divisor += 2
    return True


def ensure_prime_cache_upto(number: int) -> None:
    candidate = PRIME_CACHE[-1] + 1
    while PRIME_CACHE[-1] < number:
        if is_prime(candidate):
            PRIME_CACHE.append(candidate)
        candidate += 1


def prime_index(number: int) -> int:
    if not is_prime(number):
        return 0
    ensure_prime_cache_upto(number)
    return PRIME_CACHE.index(number) + 1


def calculate_phrase(phrase: str) -> CalculationResult:
    normalized = normalize_phrase(phrase)
    abjad_values = iter_mapped_values(normalized, ABJAD_MAP)
    total = sum(abjad_values)
    pronounced_total = sum(iter_mapped_values(normalized, PRONOUNCED_MAP))
    alif_total = sum(ALIF_MAP.get(char, 0) for char in normalized)
    saghir_total = sum(iter_mapped_values(normalized, SAGHIR_MAP))
    unique_letter_count, used_letters = compute_used_letters(normalized)
    breakdown = "+".join(str(value) for value in abjad_values)
    if breakdown:
        breakdown = f"{breakdown}={total}"
    parity = "زوج" if total % 2 == 0 else "فرد"
    letters = sum(1 for char in normalized if char in ABJAD_MAP)
    return CalculationResult(
        phrase=phrase.strip(),
        normalized_phrase=normalized,
        abjad_value=total,
        prime_index=prime_index(total),
        digit_root=compute_digit_root(total),
        abjad_sum=(total * (total + 1)) // 2,
        parity_label=parity,
        parity_order=round(total / 2),
        letter_count=letters,
        dot_count=compute_dot_count(normalized),
        unique_letter_count=unique_letter_count,
        used_letters=used_letters,
        pronounced_value=pronounced_total,
        alif_count=alif_total,
        abjad_saghir=saghir_total,
        breakdown=breakdown,
    )


def next_row_number(queryset) -> int:
    last = queryset.order_by("-row_number").values_list("row_number", flat=True).first()
    return (last or 0) + 1


def export_rows(entries: Iterable) -> list[list]:
    rows: list[list] = []
    for entry in entries:
        rows.append(
            [
                entry.row_number,
                entry.phrase,
                entry.abjad_value,
                entry.prime_index,
                entry.digit_root,
                entry.abjad_sum,
                entry.parity_label,
                entry.parity_order,
                entry.letter_count,
                entry.dot_count,
                entry.unique_letter_count,
                entry.used_letters,
                entry.pronounced_value,
                entry.alif_count,
                entry.abjad_saghir,
                entry.breakdown,
            ]
        )
    return rows


def build_csv_content(entries: Iterable) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(EXPORT_HEADERS.values())
    writer.writerows(export_rows(entries))
    return buffer.getvalue()


def build_excel_workbook(entries: Iterable, title: str) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31] or "Abjad"
    sheet.append(list(EXPORT_HEADERS.values()))
    for row in export_rows(entries):
        sheet.append(row)
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 42)
    sheet.freeze_panes = "A2"
    return workbook


def read_csv_phrases(uploaded_file) -> list[str]:
    raw = uploaded_file.read()
    for encoding in ("utf-8-sig", "utf-8", "cp1256"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    return extract_phrases_from_rows(reader)


def extract_phrases_from_rows(rows: Iterable[Iterable]) -> list[str]:
    materialized = [
        ["" if cell is None else str(cell).strip() for cell in row]
        for row in rows
        if any("" if cell is None else str(cell).strip() for cell in row)
    ]
    if not materialized:
        return []
    header = materialized[0]
    phrase_index = find_phrase_column(header)
    data_rows = materialized[1:] if phrase_index is not None else materialized
    if phrase_index is None:
        phrase_index = 1 if all(len(row) > 1 for row in data_rows) else 0
    phrases: list[str] = []
    for row in data_rows:
        if phrase_index >= len(row):
            continue
        phrase = row[phrase_index].strip()
        if phrase and phrase not in {"عبارت", "None"} and row[0] != "رديف":
            phrases.append(phrase)
    return phrases


def find_phrase_column(header: list[str]) -> int | None:
    candidates = {"عبارت", "sentence", "phrase", "متن"}
    for index, name in enumerate(header):
        if name in candidates:
            return index
    return None


def read_excel_phrases(uploaded_file) -> list[str]:
    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True, keep_vba=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        phrases = extract_phrases_from_rows(rows)
        if phrases:
            return phrases
    except Exception:
        uploaded_file.seek(0)
    return read_excel_phrases_from_archive(uploaded_file)


def read_excel_phrases_from_archive(uploaded_file) -> list[str]:
    uploaded_file.seek(0)
    with ZipFile(uploaded_file) as archive:
        shared_strings = read_shared_strings(archive)
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels_root}
        first_sheet = workbook_root.find("a:sheets", ns)[0]
        target = rel_map[first_sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]]
        rows = read_worksheet_rows(archive, f"xl/{target}", shared_strings)
        return extract_phrases_from_rows(rows)


def read_shared_strings(archive: ZipFile) -> list[str]:
    try:
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    strings: list[str] = []
    for item in root.findall(f"{namespace}si"):
        strings.append("".join(text.text or "" for text in item.iter(f"{namespace}t")))
    return strings


def read_worksheet_rows(archive: ZipFile, path: str, shared_strings: list[str]) -> list[list]:
    root = ET.fromstring(archive.read(path))
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    rows = []
    for row in root.find(f"{namespace}sheetData").findall(f"{namespace}row"):
        current = []
        for cell in row.findall(f"{namespace}c"):
            value = cell.find(f"{namespace}v")
            if value is None:
                current.append("")
                continue
            if cell.attrib.get("t") == "s":
                current.append(shared_strings[int(value.text)])
            else:
                current.append(value.text)
        rows.append(current)
    return rows


def read_phrases_from_upload(uploaded_file) -> list[str]:
    suffix = Path(uploaded_file.name).suffix.lower()
    uploaded_file.seek(0)
    if suffix == ".csv":
        return read_csv_phrases(uploaded_file)
    if suffix in {".xlsm", ".xlsx"}:
        return read_excel_phrases(uploaded_file)
    raise ValueError("فرمت فایل پشتیبانی نمی شود.")


def build_duplicate_groups(entries: Iterable) -> list[dict]:
    grouped: dict[str, list] = {}
    for entry in entries:
        key = duplicate_key(entry.phrase)
        if not key:
            continue
        grouped.setdefault(key, []).append(entry)
    results = []
    for key, items in grouped.items():
        if len(items) < 2:
            continue
        ordered = sorted(items, key=lambda item: (item.row_number, item.pk))
        results.append(
            {
                "key": key,
                "keep_entry": ordered[0],
                "entries": ordered,
                "count": len(ordered),
            }
        )
    results.sort(key=lambda group: (-group["count"], group["keep_entry"].row_number))
    return results


def export_sql_dump() -> str:
    database_path = Path(settings.DATABASES["default"]["NAME"])
    with sqlite3.connect(database_path) as connection:
        return "\n".join(connection.iterdump())


def export_sqlite_bytes() -> bytes:
    database_path = Path(settings.DATABASES["default"]["NAME"])
    return database_path.read_bytes()


def import_sqlite_file(uploaded_file) -> None:
    database_path = Path(settings.DATABASES["default"]["NAME"])
    uploaded_file.seek(0)
    data = uploaded_file.read()
    if not data.startswith(b"SQLite format 3"):
        raise ValueError("فایل sqlite معتبر نیست.")
    connections.close_all()
    database_path.write_bytes(data)


def import_sql_dump(uploaded_file) -> None:
    database_path = Path(settings.DATABASES["default"]["NAME"])
    uploaded_file.seek(0)
    raw = uploaded_file.read()
    for encoding in ("utf-8-sig", "utf-8", "cp1256"):
        try:
            sql_text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        sql_text = raw.decode("utf-8", errors="ignore")
    temp_fd, temp_name = tempfile.mkstemp(suffix=".sqlite3")
    os.close(temp_fd)
    temp_path = Path(temp_name)
    try:
        with sqlite3.connect(temp_path) as connection:
            connection.executescript(sql_text)
        connections.close_all()
        database_path.write_bytes(temp_path.read_bytes())
    finally:
        temp_path.unlink(missing_ok=True)
